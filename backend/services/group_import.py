"""解析分组表格并按姓名批量建组."""

import csv
import io

from fastapi import HTTPException, status
from sqlalchemy.orm import Session, joinedload

from backend.models import Group, GroupMember, Project, User

_HEADER_KEYWORDS_COL0 = ("组号", "group", "编号", "序号", "小组")
_HEADER_KEYWORDS_COL1 = ("组长", "leader")
_HEADER_KEYWORDS_LEGACY = ("姓名", "名字")


def _normalize_group_name(group_no: str, *, row_index: int) -> str:
    """将表格第一列组号转为系统内小组名称。"""
    label = (group_no or "").strip()
    if not label:
        return f"第{row_index}组"
    if label.endswith("组"):
        return label
    if label.isdigit():
        return f"第{label}组"
    return label


def _resolve_user_by_name(db: Session, name: str) -> User:
    name = (name or "").strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="姓名为空")
    users = db.query(User).filter(User.name == name).all()
    if len(users) == 1:
        return users[0]
    if len(users) > 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"姓名「{name}」对应多个账号，请先在用户中心消歧或使用学号建组",
        )
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=f"未找到用户：{name}",
    )


def _parse_rows_from_csv(raw: bytes) -> list[list[str]]:
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="无法识别文件编码，请使用 UTF-8 或 GBK 编码的 CSV",
        )

    sample = text[:2048]
    delimiter = ","
    if sample.count("\t") > sample.count(","):
        delimiter = "\t"
    elif sample.count(";") > sample.count(","):
        delimiter = ";"

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [[cell.strip() for cell in row] for row in reader if any(cell.strip() for cell in row)]


def _parse_rows_from_xlsx(raw: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="服务器暂不支持 xlsx，请将 Excel 另存为 CSV 后上传",
        )

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(cells):
            rows.append(cells)
    wb.close()
    return rows


def parse_grouping_rows(raw: bytes, filename: str) -> list[list[str]]:
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _parse_rows_from_xlsx(raw)
    return _parse_rows_from_csv(raw)


def _is_header_row(row: list[str]) -> bool:
    if not row:
        return False
    c0 = row[0].strip().lower()
    c1 = (row[1] if len(row) > 1 else "").strip().lower()
    if any(kw in c0 for kw in _HEADER_KEYWORDS_COL0):
        return True
    if any(kw in c1 for kw in _HEADER_KEYWORDS_COL1):
        return True
    if any(kw in c0 for kw in _HEADER_KEYWORDS_LEGACY):
        return True
    return False


def import_groups_from_spreadsheet(
    db: Session,
    project_id: int,
    raw: bytes,
    filename: str,
    *,
    replace_existing: bool = True,
) -> list[Group]:
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="关联项目不存在")

    rows = parse_grouping_rows(raw, filename)
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="表格为空")

    if _is_header_row(rows[0]):
        rows = rows[1:]

    if replace_existing:
        for group in db.query(Group).filter(Group.project_id == project_id).all():
            db.delete(group)
        db.flush()

    created: list[Group] = []
    for idx, row in enumerate(rows, start=1):
        cells = [c.strip() for c in row]
        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            continue

        if len(cells) < 2:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {idx} 行格式不正确：至少需要组号与组长两列",
            )

        group_no = cells[0]
        leader_name = cells[1]
        member_names = [name for name in cells[2:] if name]

        if not leader_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"第 {idx} 行组长姓名为空",
            )

        leader = _resolve_user_by_name(db, leader_name)
        member_users: list[User] = [leader]
        seen_ids = {leader.id}

        for member_name in member_names:
            member = _resolve_user_by_name(db, member_name)
            if member.id not in seen_ids:
                member_users.append(member)
                seen_ids.add(member.id)

        if project.group_size and len(member_users) != project.group_size:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"第 {idx} 行（{group_no or idx}）小组人数为 {len(member_users)} 人，"
                    f"项目要求 {project.group_size} 人（含组长）"
                ),
            )

        group = Group(
            name=_normalize_group_name(group_no, row_index=idx),
            project_id=project_id,
            leader_id=leader.id,
        )
        db.add(group)
        db.flush()

        for member_user in member_users:
            db.add(GroupMember(group_id=group.id, user_id=member_user.id))

        created.append(group)

    if not created:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="未解析到有效分组行")

    db.commit()

    group_ids = [g.id for g in created]
    return (
        db.query(Group)
        .options(joinedload(Group.members).joinedload(GroupMember.user))
        .filter(Group.id.in_(group_ids))
        .order_by(Group.name)
        .all()
    )
